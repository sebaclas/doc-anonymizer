"""
Manages the pseudonym mapping table: entity text → fake text.
Persists as JSON (source of truth) and optionally as Excel for user editing.
"""
import json
from pathlib import Path
from anonymizer.models import Entity

try:
    import openpyxl
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False


class AutoPseudonymGenerator:
    """
    Generates sequential pseudonyms based on entity category.
    Example: Persona1, Persona2, Org1, Org2...
    """
    PREFIX_MAP = {
        "PERSONA": "Persona",
        "PERSON": "Persona",
        "ORGANIZACIÓN": "Org",
        "ORG": "Org",
        "LUGAR": "Lugar",
        "LOC": "Lugar",
        "LOCATION": "Lugar",
        "EMAIL": "Email",
        "TELÉFONO": "Tel",
        "PHONE": "Tel",
        "DNI/NIE": "Id",
        "ID_NUMBER": "Id",
    }
    DEFAULT_PREFIX = "Ent"

    def __init__(self):
        self.counters = {}  # prefix -> count
        self.assigned = {}  # original_text -> pseudonym

    def get_pseudonym(self, text: str, entity_type: str) -> str:
        if text in self.assigned:
            return self.assigned[text]

        prefix = self.PREFIX_MAP.get(entity_type.upper(), self.DEFAULT_PREFIX)
        
        count = self.counters.get(prefix, 0) + 1
        self.counters[prefix] = count
        
        pseudo = f"{prefix}{count}"
        self.assigned[text] = pseudo
        return pseudo


def build_mapping(
    entities: list[Entity], 
    existing: dict[str, str] | None = None,
    generator: AutoPseudonymGenerator | None = None
) -> dict[str, str]:
    """
    Build a mapping dict for the given entities.
    Entities not yet in the mapping get an empty placeholder or a suggestion if generator provided.
    """
    mapping = dict(existing or {})
    for ent in entities:
        if ent.text not in mapping:
            if generator:
                mapping[ent.text] = generator.get_pseudonym(ent.text, ent.entity_type)
            else:
                mapping[ent.text] = ""
    return mapping


def load_json(path: str | Path) -> dict[str, str]:
    path = Path(path)
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(mapping: dict[str, str], path: str | Path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


def save_excel(mapping: dict[str, str], path: str | Path):
    if not _EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado.")
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapeo"

    # Headers
    ws["A1"] = "Texto original"
    ws["B1"] = "Pseudónimo"
    for cell in ws["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D5E8F0")

    for row_idx, (original, pseudo) in enumerate(mapping.items(), start=2):
        ws.cell(row=row_idx, column=1, value=original)
        ws.cell(row=row_idx, column=2, value=pseudo)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35

    wb.save(str(path))


def _create_instruction_sheet(wb: "openpyxl.Workbook"):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_properties.tabColor = "1072BA"
    
    # Simple instructions
    instructions = [
        ("Guía de Uso para Anonimización", "bold", 14),
        ("", None, 11),
        ("Columnas:", "bold", 12),
        ("Original", "bold", 11),
        ("  El texto encontrado en el documento.", None, 11),
        ("Contexto", "bold", 11),
        ("  Frase donde se encontró el texto (5 palabras antes y después).", None, 11),
        ("Tipo", "bold", 11),
        ("  La categoría (PERSONA, ORGANIZACIÓN, etc.).", None, 11),
        ("Pseudónimo", "bold", 11),
        ("  El texto por el cual se reemplazará. Puedes escribir aquí lo que quieras.", None, 11),
        ("Acción", "bold", 11),
        ("  - s : Aceptar y procesar este reemplazo.", None, 11),
        ("  - n : Ignorar. No reemplazar.", None, 11),
        ("  - vacio : idem n).", None, 11),
        ("  - Para modificar un pseudonimo detectado simplemente editalo y ponele s", None, 11),
        ("Origen", "bold", 11),
        ("  De dónde provino el hit. (DB, REGEX, NER).", None, 11),
        ("Guardar DB", "bold", 11),
        ("  Escribe 's' para guardar este par en la Base de Datos Maestra.", None, 11),
        ("", None, 11),
        ("Funciones Avanzadas (Base de Datos Maestra):", "bold", 12),
        ("Aliases", "bold", 11),
        ("  Diferentes nombres para lo mismo (ej: 'ACME, ACME Corp, ACME S.A.'). Separar por comas.", None, 11),
        ("Modo", "bold", 11),
        ("  - palabra: Búsqueda exacta (evita cambiar 'Ana' dentro de 'Mariana'). RECOMENDADO.", None, 11),
        ("  - substring: Búsqueda en cualquier parte (ideal para siglas o códigos).", None, 11),
        ("", None, 11),
        ("Importante:", "bold", 12),
        ("Solo se aplicarán los reemplazos que tengan una 's' (o 'e') en la columna 'Acción'.", "italic", 11)
    ]
    
    import openpyxl
    for row_idx, (text, weight, size) in enumerate(instructions, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = openpyxl.styles.Font(bold=(weight=="bold"), italic=(weight=="italic"), size=size)
        
    ws.column_dimensions["A"].width = 80

def save_extended_excel(data_rows: list[dict], path: str | Path):
    if not _EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado.")
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detecciones"

    headers = ["Original", "Contexto", "Pseudonimo", "Tipo", "Accion", "Guardar DB", "Origen", "Aliases", "Modo"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D5E8F0")

    green_fill = openpyxl.styles.PatternFill("solid", fgColor="C6E0B4")

    for row_idx, row_dict in enumerate(data_rows, start=2):
        ws.cell(row=row_idx, column=1, value=row_dict.get("original", ""))
        ws.cell(row=row_idx, column=2, value=row_dict.get("contexto", ""))
        ws.cell(row=row_idx, column=3, value=row_dict.get("pseudonimo", ""))
        ws.cell(row=row_idx, column=4, value=row_dict.get("tipo", ""))
        ws.cell(row=row_idx, column=5, value=row_dict.get("accion", ""))
        ws.cell(row=row_idx, column=6, value=row_dict.get("guardar_db", ""))
        
        origen = row_dict.get("origen", "")
        ws.cell(row=row_idx, column=7, value=origen)

        # New columns for DB sync
        aliases_list = row_dict.get("aliases", [])
        aliases_str = ", ".join(aliases_list) if isinstance(aliases_list, list) else str(aliases_list)
        ws.cell(row=row_idx, column=8, value=aliases_str)
        ws.cell(row=row_idx, column=9, value=row_dict.get("modo", "palabra"))
        
        # Color row if origen is DB
        if origen == "DB":
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = green_fill

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50 
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 14
    
    _create_instruction_sheet(wb)

    wb.save(str(path))


def load_extended_data(path: str | Path) -> list[dict]:
    """
    Returns full rows as list of dicts:
    [{'original': ..., 'pseudonym': ..., 'save_db': bool}, ...]
    """
    if not _EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado.")
    path = Path(path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb["Detecciones"] if "Detecciones" in wb.sheetnames else wb.active

    results = []
    # Determinamos si es el formato nuevo con Contexto o el viejo
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    has_context = "Contexto" in header_row if header_row else False
    
    idx_pseudo = 2 if has_context else 1
    idx_tipo = 3 if has_context else 2
    idx_accion = 4 if has_context else 3
    idx_save_db = 5 if has_context else 4
    idx_aliases = 7 if has_context else 6
    idx_modo = 8 if has_context else 7

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5 or not row[0]: 
            continue
            
        original = str(row[0])
        pseudo = str(row[idx_pseudo]) if row[idx_pseudo] else ""
        accion = str(row[idx_accion]).lower().strip() if row[idx_accion] else ""
        save_db = str(row[idx_save_db]).lower().strip() if row[idx_save_db] else ""
        tipo = str(row[idx_tipo]) if row[idx_tipo] else "PERSONALIZADO"

        # Optional new columns
        aliases = []
        if len(row) > idx_aliases and row[idx_aliases]:
            raw_aliases = str(row[idx_aliases])
            aliases = [a.strip() for a in raw_aliases.split(",") if a.strip()]
        
        modo = "palabra"
        if len(row) > idx_modo and row[idx_modo]:
            raw_modo = str(row[idx_modo]).strip().lower()
            if raw_modo in ["palabra", "substring"]:
                modo = raw_modo

        if accion in ["s", "e", "si", "sí"]:
            results.append({
                "original": original,
                "pseudonimo": pseudo,
                "save_db": save_db in ["s", "si", "sí", "x"],
                "tipo": tipo,
                "aliases": aliases,
                "modo": modo
            })
            
    return results


def load_extended_excel(path: str | Path) -> dict[str, str]:
    """Compatible with existing flow, return only standard mapping."""
    data = load_extended_data(path)
    return {d["original"]: d["pseudonimo"] for d in data}


def load_excel(path: str | Path) -> dict[str, str]:
    if not _EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado.")
    path = Path(path)
    wb = openpyxl.load_workbook(str(path))
    
    # Intentar buscar la hoja "Detecciones" para el nuevo formato
    if "Detecciones" in wb.sheetnames:
        return load_extended_excel(path)
        
    # Compatibilidad con formato viejo "Mapeo" o archivo simple
    if "Mapeo" in wb.sheetnames:
        ws = wb["Mapeo"]
    else:
        # Busca primera hoja que no sea instrucciones
        not_inst = [s for s in wb.sheetnames if s != "Instrucciones"]
        ws = wb[not_inst[0]] if not_inst else wb.active
    
    # Comprobar ultima opcion (quizás lo renombraron pero mantuvieron headers)
    if ws.cell(1, 4).value == "Accion":
        return load_extended_excel(path)

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        original, pseudo = row[0], row[1]
        if original:
            mapping[str(original)] = str(pseudo) if pseudo else ""
    return mapping


def validate(mapping: dict[str, str]) -> list[str]:
    """Return list of keys with empty pseudonyms."""
    return [k for k, v in mapping.items() if not v.strip()]
