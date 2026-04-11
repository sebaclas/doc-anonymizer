"""
Persistent database of known entities and their fixed pseudonyms.
Stored at ~/.doc-anonymizer/known_entities.json
"""
import json
from pathlib import Path
from dataclasses import dataclass, asdict
from anonymizer.models import EntityType
from anonymizer.config import current_settings

def _get_db_path() -> Path:
    return Path(current_settings.db_path)


@dataclass
class KnownEntity:
    original: str
    pseudonym: str
    entity_type: str  # EntityType.value string
    aliases: list[str]  # alternative spellings / abbreviations
    match_mode: str = "palabra"  # "palabra" = word boundaries (\b), "substring" = anywhere

    def to_dict(self) -> dict:
        return asdict(self)

    @staticmethod
    def from_dict(d: dict) -> "KnownEntity":
        return KnownEntity(
            original=d["original"],
            pseudonym=d["pseudonym"],
            entity_type=d.get("entity_type", EntityType.CUSTOM.value),
            aliases=d.get("aliases", []),
            match_mode=d.get("match_mode", "palabra"),
        )

    @property
    def all_forms(self) -> list[str]:
        """All text forms that should match this entity."""
        return [self.original] + self.aliases


def migrate_old_json():
    """If old JSON exists and Excel doesn't, migrate data."""
    old_json = _get_db_path().parent / "known_entities.json"
    new_xlsx = _get_db_path()
    
    if old_json.exists() and not new_xlsx.exists():
        try:
            with open(old_json, encoding="utf-8") as f:
                data = json.load(f)
            entities = [KnownEntity.from_dict(e) for e in data]
            save(entities)
            # Rename old file to avoid re-migration
            old_json.rename(old_json.with_suffix(".json.migrated"))
        except Exception:
            pass


def load(path: Path | None = None) -> list[KnownEntity]:
    """Load from Excel. If Excel is missing, try migrating from JSON."""
    db_path = path or _get_db_path()
    
    # Check for migration first
    if not db_path.exists() and path is None:
        migrate_old_json()
        
    if not db_path.exists():
        return []
        
    # We use from_excel logic internally but returning the list
    try:
        return _load_from_excel_file(db_path)
    except Exception as e:
        # Fallback/Error handling
        return []


def save(entities: list[KnownEntity], path: Path | None = None):
    """Save the entities list directly to Excel."""
    db_path = path or _get_db_path()
    _save_to_excel_file(entities, db_path)


def add(entity: KnownEntity, path: Path | None = None):
    """Add/Update entity in the Excel DB."""
    entities = load(path)
    # Replace if original already exists
    entities = [e for e in entities if e.original != entity.original]
    entities.append(entity)
    save(entities, path)


def remove(original: str, path: Path | None = None) -> bool:
    """Remove entity from the Excel DB."""
    entities = load(path)
    before = len(entities)
    entities = [e for e in entities if e.original != original]
    if len(entities) < before:
        save(entities, path)
        return True
    return False


def to_excel(excel_path: Path, db_path: Path | None = None):
    """
    Deprecated: DB is already an Excel. 
    This now acts as a 'Copy To' function for user managed file.
    """
    entities = load(db_path)
    _save_to_excel_file(entities, excel_path)
    return len(entities)


def from_excel(excel_path: Path, db_path: Path | None = None) -> tuple[int, int]:
    """
    Import/Update the master DB from an external Excel file.
    Since we are 'Excel as Master', this effectively synchronizes 
    the external file content into our persistent master file.
    """
    new_entities = _load_from_excel_file(excel_path)
    # If the user wants Excel to Rule, we replace the whole DB
    save(new_entities, db_path)
    return 0, len(new_entities) # Dummy counts for backward compat


def import_from_mapping(mapping: dict[str, str], entity_type: str = "PERSONALIZADO",
                         path: Path | None = None) -> int:
    """Bulk-import from an existing JSON mapping dict."""
    entities = load(path)
    existing = {e.original for e in entities}
    added = 0
    for original, pseudonym in mapping.items():
        if original not in existing and pseudonym:
            entities.append(KnownEntity(
                original=original,
                pseudonym=pseudonym,
                entity_type=entity_type,
                aliases=[],
            ))
            added += 1
    save(entities, path)
    return added


def _load_from_excel_file(path: Path) -> list[KnownEntity]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    entities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        original = str(row[0]).strip()
        pseudonym = str(row[1]).strip()
        entity_type = str(row[2]).strip() if row[2] else "PERSONALIZADO"
        aliases_raw = str(row[3]).strip() if row[3] else ""
        aliases = [a.strip() for a in aliases_raw.split(",") if a.strip()] if aliases_raw else []
        match_mode = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "palabra"
        if match_mode not in ("palabra", "substring"):
            match_mode = "palabra"
            
        entities.append(KnownEntity(
            original=original,
            pseudonym=pseudonym,
            entity_type=entity_type,
            aliases=aliases,
            match_mode=match_mode
        ))
    return entities


def _save_to_excel_file(entities: list[KnownEntity], path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entidades conocidas"
    
    headers = ["Original", "Pseudonimo", "Tipo", "Aliases (separados por coma)", "Modo"]
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, e in enumerate(entities, start=2):
        ws.cell(row=row_idx, column=1, value=e.original)
        ws.cell(row=row_idx, column=2, value=e.pseudonym)
        ws.cell(row=row_idx, column=3, value=e.entity_type)
        ws.cell(row=row_idx, column=4, value=", ".join(e.aliases) if e.aliases else "")
        ws.cell(row=row_idx, column=5, value=e.match_mode)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"
    
    try:
        wb.save(str(path))
    except PermissionError:
        # File is likely open in Excel
        raise PermissionError(f"No se pudo guardar en {path.name}. Por favor cerra el archivo en Excel y volvé a intentar.")
